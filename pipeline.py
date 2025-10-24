#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
pipeline.py
Flujo principal del sistema EukaryotesRegistry:
- Para cada filo (phylum), obtiene clases y especies
- Descarga sus ensamblajes genómicos
- Evalúa la calidad y selecciona los mejores por clase
- Ejecuta en paralelo con control de carga y registro
-cambio
"""
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import csv
from concurrent.futures import ThreadPoolExecutor, as_completed
from taxonomy import taxid_by_name, related_ids, names_for_ids
from genomes import genome_dataset_report_for_taxid, pick_best_assembly, extract_metrics, compute_score, passes_quality_filter
from config import MAX_WORKERS, log_kv, log_header, log_line

# ===================== FUNCIONES INTERNAS =====================


def best_species_rows_for_class(class_id: str, class_name: str, species_limit: int=20) -> list[dict]:
    """
    Obtiene las mejores especies para una clase, verificando métricas de calidad.
    """
    rows = []
    try:
        species_ids = related_ids(class_id, rank_upper="SPECIES")
        if not species_ids:
            log_kv("WARN", "Clase sin especies detectadas", Class=class_name)
            return []

        species_ids = species_ids[:species_limit]
        species_meta = {
            x["tax_id"]: x["name"]
            for x in names_for_ids(species_ids)
            if x["rank"] == "SPECIES" and x["name"]
        }

        log_kv("INFO", "Analizando clase", Class=class_name, SpeciesCount=len(species_meta))

        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = {
                executor.submit(genome_dataset_report_for_taxid, sp_id): (sp_id, sp_name)
                for sp_id, sp_name in species_meta.items()
            }

            for future in as_completed(futures):
                sp_id, sp_name = futures[future]
                try:
                    reports = future.result()
                    best = pick_best_assembly(reports)
                    if not best:
                        continue

                    metrics = extract_metrics(best)

                    # Validar cobertura y métricas completas
                    if not passes_quality_filter(metrics):
                        log_kv("WARN", "Descartado por baja cobertura", Class=class_name, Species=sp_name)
                        continue

                    for key in ["Genome coverage", "Contig N50 (kb)", "Scaffold N50 (kb)"]:
                        if metrics.get(key) is None:
                            metrics[key] = 0.0

                    metrics["Score"] = compute_score(metrics)
                    rows.append({
                        "Class": class_name,
                        "Species": sp_name,
                        **{k: metrics.get(k) for k in [
                            "Accession", "RefSeq category", "Genome level",
                            "Genome coverage", "Contig N50 (kb)", "Scaffold N50 (kb)", "Score"
                        ]}
                    })
                except Exception as e:
                    log_kv("ERROR", "Error analizando especie", Class=class_name, Species=sp_name, Error=str(e))

    except Exception as e:
        log_kv("ERROR", "Error procesando clase", Class=class_name, Error=str(e))

    # Ordenar dentro de la clase por score descendente
    rows.sort(key=lambda r: r.get("Score", 0), reverse=True)
    return rows


def best_two_species_per_class(phylum_name: str, species_per_class: int=20) -> list[dict]:
    """
    Para un filo dado:
    1. Obtiene sus clases.
    2. Evalúa sus especies.
    3. Conserva las 2 mejores especies por clase según score.
    """
    try:
        phylum_id = taxid_by_name(phylum_name)
    except Exception as e:
        log_kv("ERROR", "Filo no encontrado", Phylum=phylum_name, Error=str(e))
        return []

    log_header(f"Procesando filo {phylum_name}")
    class_ids = related_ids(phylum_id, rank_upper="CLASS")
    class_info = {
        x["tax_id"]: x["name"]
        for x in names_for_ids(class_ids)
        if x["rank"] == "CLASS" and x["name"]
    }

    all_rows = []

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {
            executor.submit(best_species_rows_for_class, cid, cname, species_per_class): (cid, cname)
            for cid, cname in class_info.items()
        }

        for future in as_completed(futures):
            cid, cname = futures[future]
            try:
                rows = future.result()
                if not rows:
                    continue

                # Ordenar por calidad si existe cobertura o N50
                def sort_key(r):
                    cov = r.get("Genome coverage") or 0
                    n50 = r.get("Scaffold N50 (kb)") or 0
                    return (cov, n50)

                top2 = sorted(rows, key=lambda r: r.get("Score", 0), reverse=True)[:2]
                for r in top2:
                    r["Phylum"] = phylum_name
                    r["Class"] = cname
                all_rows.extend(top2)

                log_kv("INFO", "Clase procesada", Phylum=phylum_name, Class=cname, Selected=len(top2))
            except Exception as e:
                log_kv("ERROR", "Error en clase", Class=cname, Error=str(e))

    log_kv("INFO", "Filo completado", Phylum=phylum_name, Rows=len(all_rows))
    return all_rows


def export_results(rows: list[dict]):
    """
    Exporta los resultados a Excel con formato profesional:
    - Fondo amarillo en encabezados
    - Bordes visibles
    - Autoajuste de columnas
    - Cada Phylum en una hoja separada
    """
    if not rows:
        log_kv("WARN", "Sin resultados que exportar.")
        return

    phyla = sorted(set(r.get("Phylum", "Unknown") for r in rows))
    output_excel = "best_genomes_by_class.xlsx"

    for phylum in phyla:
        subset = [r for r in rows if r.get("Phylum") == phylum]
        if not subset:
            log_kv("WARN", f"Sin datos para {phylum}")
            continue

        df = pd.DataFrame(subset)
        df.sort_values(by=["Phylum", "Class", "Species"], inplace=True, kind="stable")


        # Guardar hoja (crear o anexar)
        if os.path.exists(output_excel):
            mode = "a"
            writer_args = {"if_sheet_exists": "replace"}
        else:
            mode = "w"
            writer_args = {}


        try:
            
            with pd.ExcelWriter(output_excel, mode=mode, engine="openpyxl", **writer_args) as writer:
                sheet_name = phylum[:30]
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            log_kv("ERROR", f"Fallo al escribir hoja de {phylum}: {e}")
            continue

        # Aplicar formato visual
        try:
            wb = load_workbook(output_excel)
            ws = wb[sheet_name]

            # Estilos
            header_fill = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
            header_font = Font(bold=True)
            border = Border(
                left=Side(style="thin", color="000000"),
                right=Side(style="thin", color="000000"),
                top=Side(style="thin", color="000000"),
                bottom=Side(style="thin", color="000000"),
            )

            # Encabezado
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            # Celdas con bordes y texto centrado
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Ajuste automático del ancho
            for col in ws.columns:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 3, 45)

            wb.save(output_excel)
            wb.close()
            log_kv("INFO", f"Hoja '{phylum}' exportada ({len(df)} filas)")
        except Exception as e:
            log_kv("ERROR", f"Fallo al formatear hoja '{phylum}': {e}")

    log_kv("INFO", f"Exportación finalizada -> {output_excel}")

